from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any, Iterable

from wq_pipeline.core.geo import coordinate_from_record
from wq_pipeline.core.records import as_float, safe_string
from wq_pipeline.core.status import make_source_status


SOURCE_ID = "epa_official_chemistry"
CHEMISTRY_DIR = Path(__file__).resolve().parents[3] / "data" / "source" / "epa-chemistry"

DATE_KEYS = [
    "sample_date", "sampling_date", "date", "result_date", "monitoring_date",
    "sampledate", "samplingdate", "sample_taken", "sample_taken_date",
]
STATION_CODE_KEYS = [
    "station_code", "monitoring_station_code", "stationcode", "station",
    "location_code", "site_code", "sample_point_code",
]
STATION_NAME_KEYS = [
    "station_name", "monitoring_station_name", "stationname", "location",
    "location_name", "site_name", "sample_point_name",
]
WATERBODY_CODE_KEYS = [
    "waterbody_code", "water_body_code", "waterbodycode", "wb_code",
    "water_body_eu_code", "eu_code", "waterbody_eu_code",
]
WATERBODY_NAME_KEYS = [
    "waterbody_name", "water_body_name", "waterbodyname", "wb_name",
]
PARAMETER_KEYS = [
    "determinand", "determinand_label", "determinand_name", "parameter",
    "parameter_name", "analyte", "chemical", "variable", "test",
]
VALUE_KEYS = [
    "result", "result_value", "value", "concentration", "numeric_value",
    "measurement", "observed_value",
]
UNIT_KEYS = ["unit", "units", "result_unit", "uom", "measurement_unit"]
LAT_KEYS = ["lat", "latitude"]
LON_KEYS = ["lon", "lng", "long", "longitude"]
EASTING_KEYS = ["easting", "east", "x"]
NORTHING_KEYS = ["northing", "north", "y"]

KNOWN_WIDE_PARAMETERS = {
    "orthophosphate": "Orthophosphate",
    "orthophosphate_as_p": "Orthophosphate as P",
    "ortho_phosphate": "Orthophosphate",
    "molybdate_reactive_phosphorus": "Molybdate reactive phosphorus",
    "mrp": "Molybdate reactive phosphorus",
    "phosphate": "Phosphate",
    "po4": "Phosphate",
    "po4_p": "Orthophosphate as P",
    "total_phosphorus": "Total phosphorus",
    "tp": "Total phosphorus",
    "nitrate": "Nitrate",
    "nitrate_as_n": "Nitrate as N",
    "no3": "Nitrate",
    "no3_n": "Nitrate as N",
    "nitrite": "Nitrite",
    "nitrite_as_n": "Nitrite as N",
    "no2": "Nitrite",
    "no2_n": "Nitrite as N",
    "ammonia": "Ammonia",
    "ammoniacal_nitrogen": "Ammoniacal nitrogen",
    "ammonium": "Ammonium",
    "nh4": "Ammonium",
    "nh4_n": "Ammoniacal nitrogen as N",
    "total_oxidised_nitrogen": "Total oxidised nitrogen",
    "ton": "Total oxidised nitrogen",
    "ton_n": "Total oxidised nitrogen as N",
    "total_nitrogen": "Total nitrogen",
    "tn": "Total nitrogen",
    "bod": "Biochemical oxygen demand",
    "dissolved_oxygen": "Dissolved oxygen",
    "do": "Dissolved oxygen",
    "conductivity": "Conductivity",
    "ph": "pH",
    "temperature": "Temperature",
    "e_coli": "E. coli",
    "intestinal_enterococci": "Intestinal enterococci",
}


def normalise_key(value: Any) -> str:
    return re.sub(r"_+", "_", re.sub(r"[^a-z0-9]+", "_", safe_string(value, "").lower())).strip("_")


def clean_text(value: Any) -> str:
    return safe_string(value, "").strip()


def parse_number(value: Any) -> float | None:
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = clean_text(value)
    if not text:
        return None

    text = text.replace("\u00a0", " ").strip()
    text = re.sub(r"^[<>=~ ]+", "", text)
    text = text.replace(" ", "")

    if "," in text and "." not in text:
        text = text.replace(",", "" if text.count(",") > 1 else ".")

    try:
        return float(text)
    except ValueError:
        return None


def find(row: dict[str, Any], keys: Iterable[str]) -> Any:
    for key in keys:
        norm = normalise_key(key)
        if norm in row and row[norm] not in (None, ""):
            return row[norm]
    return ""


def result_unit(row: dict[str, Any], parameter_key: str = "") -> str:
    explicit = clean_text(find(row, UNIT_KEYS))

    if explicit:
        return explicit

    if parameter_key in {"ph"}:
        return "pH units"

    return ""


def file_paths(directory: Path = CHEMISTRY_DIR) -> list[Path]:
    if not directory.exists():
        return []

    allowed = {".csv", ".tsv", ".txt", ".xlsx"}
    return sorted(
        path for path in directory.iterdir()
        if path.is_file()
        and path.suffix.lower() in allowed
        and not path.name.startswith("~$")
    )


def normalise_row(row: dict[str, Any]) -> dict[str, Any]:
    return {normalise_key(key): value for key, value in row.items() if normalise_key(key)}


def read_csv_rows(path: Path) -> list[dict[str, Any]]:
    raw = path.read_text(encoding="utf-8-sig", errors="replace")
    sample = raw[:4096]

    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
        if path.suffix.lower() == ".tsv":
            dialect.delimiter = "\t"

    reader = csv.DictReader(raw.splitlines(), dialect=dialect)
    return [normalise_row(row) for row in reader if any(clean_text(value) for value in row.values())]


def read_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return []

    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []

    for ws in wb.worksheets:
        iterator = ws.iter_rows(values_only=True)
        header = None

        for raw_row in iterator:
            values = list(raw_row or [])

            if header is None:
                if not any(value is not None and clean_text(value) for value in values):
                    continue

                header = [normalise_key(value) for value in values]
                continue

            if not any(value is not None and clean_text(value) for value in values):
                continue

            row = {key: values[index] if index < len(values) else "" for index, key in enumerate(header) if key}
            row["_sheet"] = ws.title
            rows.append(row)

    return rows


def read_rows(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".xlsx":
        return read_xlsx_rows(path)

    return read_csv_rows(path)


def parameter_candidates(row: dict[str, Any]) -> list[tuple[str, str, float, str]]:
    long_parameter = clean_text(find(row, PARAMETER_KEYS))
    long_value = parse_number(find(row, VALUE_KEYS))

    if long_parameter and long_value is not None:
        key = normalise_key(long_parameter)
        return [(key, long_parameter, long_value, result_unit(row, key))]

    candidates: list[tuple[str, str, float, str]] = []

    for key, label in KNOWN_WIDE_PARAMETERS.items():
        if key not in row:
            continue

        value = parse_number(row[key])

        if value is not None:
            candidates.append((key, label, value, result_unit(row, key)))

    return candidates


def build_context_index(context_records: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    index: dict[str, dict[str, Any]] = {}

    for record in context_records:
        if record.get("source") != "epa_official_wq":
            continue

        params = {
            normalise_key(parameter.get("key")): parameter.get("value")
            for parameter in record.get("parameters", [])
            if isinstance(parameter, dict)
        }

        info = {
            "lat": record.get("lat"),
            "lon": record.get("lon"),
            "focus_area_ids": record.get("focus_area_ids", []),
            "station_code": params.get("station_code", ""),
            "station_name": params.get("station_name", ""),
            "waterbody_code": params.get("waterbody_code", ""),
            "waterbody_name": params.get("waterbody_name", ""),
        }

        for value in (
            info["station_code"],
            info["station_name"],
            info["waterbody_code"],
            info["waterbody_name"],
            record.get("name"),
        ):
            key = normalise_key(value)
            if key:
                index[key] = info

    return index


def context_match(row: dict[str, Any], index: dict[str, dict[str, Any]]) -> dict[str, Any]:
    for value in (
        find(row, STATION_CODE_KEYS),
        find(row, STATION_NAME_KEYS),
        find(row, WATERBODY_CODE_KEYS),
        find(row, WATERBODY_NAME_KEYS),
    ):
        key = normalise_key(value)
        if key and key in index:
            return index[key]

    return {}


def record_id(path: Path, row_number: int, parameter_key: str, station_code: str, observed_at: str) -> str:
    bits = [
        path.stem,
        str(row_number),
        parameter_key,
        normalise_key(station_code),
        normalise_key(observed_at),
    ]
    return "official-chemistry:" + ":".join(bit for bit in bits if bit)


def build_records_from_file(
    path: Path,
    *,
    now: str,
    source_defs: dict[str, dict[str, Any]],
    context_index: dict[str, dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[str]]:
    records: list[dict[str, Any]] = []
    errors: list[str] = []

    try:
        rows = read_rows(path)
    except Exception as exc:
        return [], [f"{path.name}: {exc}"]

    for row_number, row in enumerate(rows, start=2):
        match = context_match(row, context_index)

        observed_at = clean_text(find(row, DATE_KEYS))
        station_code = clean_text(find(row, STATION_CODE_KEYS)) or clean_text(match.get("station_code"))
        station_name = clean_text(find(row, STATION_NAME_KEYS)) or clean_text(match.get("station_name"))
        waterbody_code = clean_text(find(row, WATERBODY_CODE_KEYS)) or clean_text(match.get("waterbody_code"))
        waterbody_name = clean_text(find(row, WATERBODY_NAME_KEYS)) or clean_text(match.get("waterbody_name"))

        base_record = {
            "lat": find(row, LAT_KEYS),
            "lon": find(row, LON_KEYS),
            "Easting": find(row, EASTING_KEYS),
            "Northing": find(row, NORTHING_KEYS),
        }

        lat, lon = coordinate_from_record(base_record, f"{station_name} {waterbody_name}")
        lat = lat if lat is not None else match.get("lat")
        lon = lon if lon is not None else match.get("lon")

        for parameter_key, label, value, unit in parameter_candidates(row):
            name_parts = [label]
            if station_name:
                name_parts.append(station_name)
            elif waterbody_name:
                name_parts.append(waterbody_name)

            records.append({
                "id": record_id(path, row_number, parameter_key, station_code or waterbody_code, observed_at),
                "source": SOURCE_ID,
                "source_label": source_defs[SOURCE_ID]["name"],
                "type": "official_chemistry_result",
                "freshness": "official_historic",
                "name": "Official chemistry — " + " · ".join(name_parts),
                "lat": lat,
                "lon": lon,
                "observed_at": observed_at or None,
                "generated_at": now,
                "status": f"{label}: {value:g}{(' ' + unit) if unit else ''}",
                "description": "Official chemistry result imported from a Catchments.ie/EPA chemistry download file.",
                "url": "https://www.catchments.ie/data/",
                "focus_area_ids": list(match.get("focus_area_ids", [])),
                "parameters": [
                    {
                        "key": parameter_key,
                        "label": label,
                        "value": value,
                        "unit": unit,
                        "basis": f"Official chemistry download: {path.name}",
                    },
                    {
                        "key": "sample_date",
                        "label": "Sample date",
                        "value": observed_at,
                        "unit": "",
                        "basis": f"Official chemistry download: {path.name}",
                    },
                    {
                        "key": "station_code",
                        "label": "Station code",
                        "value": station_code,
                        "unit": "",
                        "basis": f"Official chemistry download: {path.name}",
                    },
                    {
                        "key": "station_name",
                        "label": "Station name",
                        "value": station_name,
                        "unit": "",
                        "basis": f"Official chemistry download: {path.name}",
                    },
                    {
                        "key": "waterbody_code",
                        "label": "Waterbody code",
                        "value": waterbody_code,
                        "unit": "",
                        "basis": f"Official chemistry download: {path.name}",
                    },
                    {
                        "key": "waterbody_name",
                        "label": "Waterbody name",
                        "value": waterbody_name,
                        "unit": "",
                        "basis": f"Official chemistry download: {path.name}",
                    },
                    {
                        "key": "source_file",
                        "label": "Source file",
                        "value": path.name,
                        "unit": "",
                        "basis": "Local official download file",
                    },
                ],
                "raw": row,
            })

    return records, errors


def harvest_official_chemistry(
    now: str,
    context_records: list[dict[str, Any]],
    *,
    source_defs: dict[str, dict[str, Any]],
    directory: Path = CHEMISTRY_DIR,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    paths = file_paths(directory)
    context_index = build_context_index(context_records)
    records: list[dict[str, Any]] = []
    errors: list[str] = []

    for path in paths:
        file_records, file_errors = build_records_from_file(
            path,
            now=now,
            source_defs=source_defs,
            context_index=context_index,
        )
        records.extend(file_records)
        errors.extend(file_errors)

    if records and errors:
        status = "partial"
    elif records:
        status = "ok"
    elif paths:
        status = "empty"
    else:
        status = "empty"

    source = make_source_status(
        source_defs,
        SOURCE_ID,
        status=status,
        records=len(records),
        fetched_at_utc=now,
        error="; ".join(errors[:4]) if errors and not records else None,
        elapsed_ms=None,
    )
    source["files"] = len(paths)
    try:
        source["directory"] = str(directory.resolve().relative_to(Path(__file__).resolve().parents[3]))
    except ValueError:
        source["directory"] = str(directory)

    return records, source
