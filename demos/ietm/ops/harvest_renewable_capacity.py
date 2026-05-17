#!/usr/bin/env python3
"""
Harvest renewable capacity evidence for the Ireland Energy Transition Monitor.

The output is deliberately conservative:
- connected / energised capacity is treated separately from contracted / planned capacity
- MEC / offer quantities are not mixed with live generation
- decommissioning is inferred only as a "removed_from_register_candidate" unless an
  official source explicitly says a unit is closed, withdrawn, terminated or retired.

Primary public evidence sources:
- EirGrid TSO connected / contracted generator pages
- ESB Networks DSO generator statistics pages
- Government RESS / ORESS pages
- EirGrid AIRAA page for official adequacy assumptions
"""

from __future__ import annotations

import hashlib
import io
import json
import re
import sys
import tempfile
from dataclasses import dataclass
from datetime import datetime, timezone
from html.parser import HTMLParser
from pathlib import Path
from typing import Any
from urllib.parse import urljoin
from urllib.request import Request, urlopen

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover - handled at runtime
    PdfReader = None  # type: ignore

try:
    import openpyxl
except Exception:  # pragma: no cover - handled at runtime
    openpyxl = None  # type: ignore


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "data" / "source"
OUT = SOURCE_DIR / "renewable_capacity.json"
AUDIT_OUT = SOURCE_DIR / "renewable_capacity_audit.json"

USER_AGENT = (
    "Mozilla/5.0 (compatible; IrelandEnergyTransitionMonitor/1.0; "
    "+https://salmonofdoubt.github.io/demos/ietm/)"
)

SOURCE_PAGES = {
    "eirgrid_system_renewable": "https://www.eirgrid.ie/grid/system-and-renewable-data-reports",
    "eirgrid_connected_contracted": "https://www.eirgrid.ie/industry/customer-information/connected-and-contracted-generators",
    "esb_generator_statistics": "https://www.esbnetworks.ie/services/get-connected/renewable-connection/generator-statistics",
    "gov_ress": "https://www.gov.ie/en/department-of-climate-energy-and-the-environment/publications/renewable-electricity-support-scheme-ress/",
    "gov_oress": "https://www.gov.ie/en/department-of-climate-energy-and-the-environment/publications/offshore-renewable-electricity-support-scheme-oress/",
    "eirgrid_airaa": "https://www.eirgrid.ie/airaa",
}

RENEWABLE_OR_STORAGE_WORDS = (
    "wind",
    "solar",
    "pv",
    "battery",
    "bess",
    "storage",
    "hydro",
    "wave",
    "biomass",
    "biogas",
    "lfg",
    "waste to energy",
)

EXCLUDE_TECH_WORDS = (
    "gas",
    "ocgt",
    "ccgt",
    "chp",
    "interconnector",
    "synchronous condenser",
    "rotating stabiliser",
    "autoproducer demand",
)


@dataclass
class Link:
    text: str
    href: str


class LinkCollector(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.links: list[Link] = []
        self._href: str | None = None
        self._chunks: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() != "a":
            return
        href = dict(attrs).get("href")
        if href:
            self._href = href
            self._chunks = []

    def handle_data(self, data: str) -> None:
        if self._href:
            self._chunks.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() == "a" and self._href:
            text = " ".join(" ".join(self._chunks).split())
            self.links.append(Link(text=text, href=self._href))
            self._href = None
            self._chunks = []


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def fetch_bytes(url: str, timeout: int = 45) -> bytes:
    req = Request(url, headers={"User-Agent": USER_AGENT})
    with urlopen(req, timeout=timeout) as response:
        return response.read()


def fetch_text(url: str) -> str:
    return fetch_bytes(url).decode("utf-8", errors="replace")


def collect_links(page_url: str) -> list[Link]:
    html = fetch_text(page_url)
    parser = LinkCollector()
    parser.feed(html)
    links: list[Link] = []
    for link in parser.links:
        links.append(Link(text=link.text, href=urljoin(page_url, link.href)))
    return links


def link_matches(link: Link, *needles: str) -> bool:
    haystack = f"{link.text} {link.href}".lower()
    return all(needle.lower() in haystack for needle in needles)


def first_matching_link(links: list[Link], *needles: str) -> str | None:
    for link in links:
        if link_matches(link, *needles):
            return link.href
    return None


def source_id(*parts: str) -> str:
    h = hashlib.sha1("|".join(parts).encode("utf-8")).hexdigest()[:12]
    return h


def canonical_project_id(name: str, network: str, status: str, basis: str) -> str:
    stem = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    if not stem:
        stem = "unnamed"
    return f"{network.lower()}-{status.lower()}-{basis.lower()}-{stem}-{source_id(name, network, status, basis)}"


def normalise_space(value: str) -> str:
    return " ".join(str(value or "").replace("\xa0", " ").split())


def infer_technology(text: str) -> str | None:
    s = normalise_space(text).lower()
    for forbidden in EXCLUDE_TECH_WORDS:
        if forbidden in s and not any(allowed in s for allowed in RENEWABLE_OR_STORAGE_WORDS):
            return None

    if "wind + battery" in s or "wind battery" in s:
        return "hybrid_renewable_storage"
    if "solar + battery" in s or "solar battery" in s:
        return "hybrid_renewable_storage"
    if "wind + solar" in s or "wind and solar" in s:
        return "hybrid_renewable"
    if "waste to energy" in s:
        return "waste_to_energy"
    if "energy storage" in s or "battery" in s or "bess" in s:
        return "battery_storage"
    if "solar" in s or re.search(r"\bpv\b", s):
        return "solar"
    if "wind" in s:
        return "wind_onshore"
    if "hydro" in s:
        return "hydro"
    if "wave" in s:
        return "wave"
    if "biomass" in s or "biogas" in s or "ad " in f"{s} " or "lfg" in s:
        return "bioenergy"
    return None


def status_from_source(source_key: str) -> str:
    if "energised" in source_key:
        return "connected"
    if "connected" in source_key:
        return "connected"
    if "contracted" in source_key:
        return "contracted"
    if "ress" in source_key or "oress" in source_key:
        return "awarded_support"
    return "evidence_observed"


def confidence_for(status: str, source_key: str) -> str:
    if status in {"connected", "contracted"} and ("eirgrid" in source_key or "esb" in source_key):
        return "high"
    if status == "awarded_support":
        return "medium_high"
    return "medium"


def parse_pdf_text(url: str) -> tuple[str, dict[str, Any]]:
    if PdfReader is None:
        raise RuntimeError("pypdf is not installed. Add pypdf to ops/requirements.txt")

    raw = fetch_bytes(url)
    reader = PdfReader(io.BytesIO(raw))
    page_texts: list[str] = []
    for page in reader.pages:
        page_texts.append(page.extract_text() or "")

    return "\n".join(page_texts), {
        "url": url,
        "bytes": len(raw),
        "pages": len(reader.pages),
    }


def parse_info_correct(text: str) -> str | None:
    patterns = [
        r"Correct as of\s+([0-9]{1,2}/[0-9]{1,2}/[0-9]{4})",
        r"Information correct:\s+([0-9]{1,2}/[0-9]{1,2}/[0-9]{4})",
        r"Correct as of\s+([0-9]{1,2}-[0-9]{1,2}-[0-9]{4})",
    ]
    for pattern in patterns:
        m = re.search(pattern, text, flags=re.I)
        if m:
            return m.group(1)
    return None


def clean_project_name(name: str) -> str:
    name = normalise_space(name)
    name = re.sub(r"\b(?:ECP|Gate|Non GPA|T-\d|PCI|LCIS|PMOD)\b.*$", "", name, flags=re.I).strip()
    name = re.sub(r"\s+(?:Wind|Solar|Battery|Biomass|Biogas|Hydro|Wave|LFG|Waste to Energy)$", "", name, flags=re.I).strip()
    return name or "Unnamed project"


def find_technology_capacity(raw: str) -> tuple[str, float, int] | None:
    """Find the technology field that is directly followed by a capacity.

    Project names often contain words such as "Solar Farm" or "Wind Farm"; this
    function avoids taking those as the technology column by requiring the
    technology label to be followed immediately by a numeric MEC/capacity value.
    """
    patterns: list[tuple[str, str]] = [
        ("hybrid_renewable_storage", r"\b(?:Solar\s*\+\s*Battery|Wind\s*\+\s*Battery)\s+(?P<cap>[0-9]{1,4}(?:\.[0-9]+)?)\b"),
        ("hybrid_renewable", r"\bWind\s*\+\s*Solar\s+(?P<cap>[0-9]{1,4}(?:\.[0-9]+)?)\b"),
        ("waste_to_energy", r"\bWaste\s+to\s+Energy\s+(?P<cap>[0-9]{1,4}(?:\.[0-9]+)?)\b"),
        ("battery_storage", r"\b(?:Battery\s+Storage|Energy\s+Storage|Battery|BESS)\s+(?P<cap>[0-9]{1,4}(?:\.[0-9]+)?)\b"),
        ("solar", r"\b(?:Solar|PV)\s+(?P<cap>[0-9]{1,4}(?:\.[0-9]+)?)\b"),
        ("wind_onshore", r"\bWind\s+(?P<cap>[0-9]{1,4}(?:\.[0-9]+)?)\b"),
        ("hydro", r"\bHydro\s+(?P<cap>[0-9]{1,4}(?:\.[0-9]+)?)\b"),
        ("wave", r"\bWave\s+(?P<cap>[0-9]{1,4}(?:\.[0-9]+)?)\b"),
        ("bioenergy", r"\b(?:Biomass|Biogas|LFG|AD)\s+(?P<cap>[0-9]{1,4}(?:\.[0-9]+)?)\b"),
    ]

    matches: list[tuple[int, str, float]] = []
    for tech, pattern in patterns:
        for match in re.finditer(pattern, raw, flags=re.I):
            try:
                capacity = float(match.group("cap"))
            except Exception:
                continue
            if capacity > 0:
                matches.append((match.start(), tech, capacity))

    if not matches:
        return None

    pos, tech, capacity = sorted(matches, key=lambda item: item[0])[-1]
    return tech, capacity, pos


def record_from_line(
    line: str,
    *,
    source_key: str,
    network: str,
    publisher: str,
    source_url: str,
    capacity_basis: str,
    source_date: str | None,
) -> dict[str, Any] | None:
    raw = normalise_space(line)
    if not raw:
        return None

    if not any(word in raw.lower() for word in RENEWABLE_OR_STORAGE_WORDS):
        return None

    if not re.search(r"\b(?:TG|DG)\d+[A-Za-z]?\b", raw):
        return None

    found = find_technology_capacity(raw)
    if found is None:
        return None

    tech, capacity_mw, tech_pos = found

    ref_match = re.search(r"\b(?P<ref>(?:TG|DG)\d+[A-Za-z]?)\b", raw)
    ref = ref_match.group("ref") if ref_match else None

    before_tech = raw[:tech_pos].strip() if tech_pos >= 0 else raw
    if ref_match and ref_match.start() < tech_pos:
        name_part = raw[ref_match.end():tech_pos].strip()
    else:
        m = re.search(r"\b(?:TG|DG)\d+[A-Za-z]?\b\s+(.*?)\s+(?:ECP|Gate|Non GPA|T-\d|PCI|LCIS|PMOD|Wind|Solar|Battery|Biomass|Hydro|Wave)", raw, flags=re.I)
        name_part = m.group(1) if m else before_tech

    name_part = re.sub(r"^[A-Za-z\s/().-]+\s+\d{2,3}\s*kV\s+", "", name_part, flags=re.I)
    name = clean_project_name(name_part)

    status = status_from_source(source_key)
    return {
        "id": canonical_project_id(name, network, status, capacity_basis),
        "name": name,
        "technology": tech,
        "network": network,
        "status": status,
        "capacity_mw": capacity_mw,
        "capacity_basis": capacity_basis,
        "confidence": confidence_for(status, source_key),
        "source_key": source_key,
        "source_correct_as_of": source_date,
        "reference": ref,
        "raw_line": raw,
        "sources": [
            {
                "publisher": publisher,
                "source_type": source_key.replace("_", " "),
                "url": source_url,
                "date_accessed": now_iso(),
            }
        ],
    }


def parse_generator_pdf(
    url: str,
    *,
    source_key: str,
    network: str,
    publisher: str,
    capacity_basis: str = "MEC MW",
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    text, meta = parse_pdf_text(url)
    source_date = parse_info_correct(text)

    lines = [normalise_space(line) for line in text.splitlines()]
    records: list[dict[str, Any]] = []

    windows: list[str] = []
    for i in range(len(lines)):
        windows.append(lines[i])
        if i + 1 < len(lines):
            windows.append(f"{lines[i]} {lines[i + 1]}")
        if i + 2 < len(lines):
            windows.append(f"{lines[i]} {lines[i + 1]} {lines[i + 2]}")

    seen_raw: set[str] = set()
    for line in windows:
        if line in seen_raw:
            continue
        seen_raw.add(line)
        rec = record_from_line(
            line,
            source_key=source_key,
            network=network,
            publisher=publisher,
            source_url=url,
            capacity_basis=capacity_basis,
            source_date=source_date,
        )
        if rec:
            records.append(rec)

    deduped: dict[str, dict[str, Any]] = {}
    for rec in records:
        key = rec["id"]
        if key not in deduped or len(rec["raw_line"]) < len(deduped[key].get("raw_line", "")):
            deduped[key] = rec

    meta.update(
        {
            "source_key": source_key,
            "records": len(deduped),
            "source_correct_as_of": source_date,
            "parser": "pdf_line_window_v1",
        }
    )
    return list(deduped.values()), meta


def scan_workbook_for_capacity(url: str, source_key: str) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Conservative workbook scanner for obvious installed-capacity aggregate rows."""
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed")

    raw = fetch_bytes(url)
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
        tmp.write(raw)
        tmp.flush()
        wb = openpyxl.load_workbook(tmp.name, data_only=True, read_only=True)

        aggregates: list[dict[str, Any]] = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                values = [normalise_space(str(v)) for v in row if v is not None]
                joined = " | ".join(values)
                low = joined.lower()
                if "installed" not in low or not any(t in low for t in ("wind", "solar", "renewable")):
                    continue
                nums = []
                for v in row:
                    try:
                        if isinstance(v, (int, float)):
                            nums.append(float(v))
                    except Exception:
                        pass
                if not nums:
                    continue
                tech = infer_technology(joined) or "renewable_mixed"
                capacity = max(nums)
                aggregates.append(
                    {
                        "id": canonical_project_id(f"{ws.title} aggregate installed {tech}", "system", "aggregate", "installed_capacity"),
                        "name": f"{ws.title} aggregate installed {tech}",
                        "technology": tech,
                        "network": "system",
                        "status": "aggregate_observed",
                        "capacity_mw": capacity,
                        "capacity_basis": "installed capacity MW inferred from workbook row",
                        "confidence": "medium",
                        "source_key": source_key,
                        "raw_line": joined[:500],
                        "sources": [
                            {
                                "publisher": "EirGrid",
                                "source_type": source_key.replace("_", " "),
                                "url": url,
                                "date_accessed": now_iso(),
                            }
                        ],
                    }
                )
        return aggregates, {
            "url": url,
            "bytes": len(raw),
            "worksheets": wb.sheetnames,
            "records": len(aggregates),
            "parser": "xlsx_aggregate_scan_v1",
            "source_key": source_key,
        }


def summarise(records: list[dict[str, Any]]) -> dict[str, Any]:
    buckets: dict[str, dict[str, float]] = {
        "connected": {},
        "contracted": {},
        "awarded_support": {},
        "aggregate_observed": {},
        "removed_from_register_candidate": {},
    }

    for rec in records:
        status = rec.get("status") or "unknown"
        tech = rec.get("technology") or "unknown"
        cap = rec.get("capacity_mw")
        try:
            cap_f = float(cap)
        except Exception:
            continue
        buckets.setdefault(status, {})
        buckets[status][tech] = round(buckets[status].get(tech, 0.0) + cap_f, 3)

    totals = {
        status: round(sum(techs.values()), 3)
        for status, techs in buckets.items()
    }
    return {
        "by_status_and_technology_mw": buckets,
        "totals_by_status_mw": totals,
        "record_count": len(records),
        "warning": (
            "MEC MW, installed capacity, RESS/ORESS offer quantity and capacity-market "
            "de-rated MW are different bases. Do not add them without qualification."
        ),
    }


def load_previous() -> dict[str, Any]:
    if not OUT.exists():
        return {}
    try:
        return json.loads(OUT.read_text())
    except Exception:
        return {}


def removed_candidates(previous: dict[str, Any], current_records: list[dict[str, Any]], successful_source_keys: set[str]) -> list[dict[str, Any]]:
    previous_records = previous.get("projects", []) or []
    current_ids = {rec.get("id") for rec in current_records}

    candidates: list[dict[str, Any]] = []
    for old in previous_records:
        old_source_key = old.get("source_key")
        old_status = old.get("status")
        if old_status == "removed_from_register_candidate":
            continue
        if old_status not in {"connected", "contracted"}:
            continue
        if old_source_key not in successful_source_keys:
            continue
        if old.get("id") in current_ids:
            continue

        rec = dict(old)
        rec["previous_status"] = old_status
        rec["status"] = "removed_from_register_candidate"
        rec["confidence"] = "low"
        rec["removal_detected_at"] = now_iso()
        rec["caveat"] = (
            "Previously present in an official connected/contracted register but absent "
            "from the latest successfully harvested version. This is not proof of decommissioning."
        )
        candidates.append(rec)
    return candidates


def harvest() -> tuple[dict[str, Any], dict[str, Any]]:
    SOURCE_DIR.mkdir(parents=True, exist_ok=True)

    audit: dict[str, Any] = {
        "generated_at": now_iso(),
        "sources": {},
        "errors": [],
    }
    records: list[dict[str, Any]] = []
    successful_source_keys: set[str] = set()

    links_by_page: dict[str, list[Link]] = {}
    for key, page in SOURCE_PAGES.items():
        try:
            links = collect_links(page)
            links_by_page[key] = links
            audit["sources"][key] = {
                "url": page,
                "link_count": len(links),
                "status": "loaded",
            }
        except Exception as exc:
            audit["errors"].append({"source": key, "url": page, "error": str(exc)})
            links_by_page[key] = []
            audit["sources"][key] = {
                "url": page,
                "status": "failed",
                "error": str(exc),
            }

    eirgrid_links = links_by_page.get("eirgrid_connected_contracted", [])
    tso_wind_links = [link for link in eirgrid_links if link_matches(link, "Contracted", "TSO", "Wind") and not link_matches(link, "Non", "Wind")]

    jobs: list[tuple[str, str | None, str, str, str, str]] = [
        (
            "eirgrid_tso_contracted_wind",
            tso_wind_links[0].href if tso_wind_links else None,
            "TSO",
            "EirGrid",
            "MEC MW",
            "pdf",
        ),
        (
            "eirgrid_tso_contracted_non_wind",
            first_matching_link(eirgrid_links, "Contracted", "TSO", "Non", "Wind"),
            "TSO",
            "EirGrid",
            "MEC MW",
            "pdf",
        ),
        (
            "esb_dso_contracted_renewable",
            first_matching_link(links_by_page.get("esb_generator_statistics", []), "DSO", "Contracted", "Renewable"),
            "DSO",
            "ESB Networks",
            "MEC MW",
            "pdf",
        ),
        (
            "esb_dso_energised_renewable",
            first_matching_link(links_by_page.get("esb_generator_statistics", []), "DSO", "Energised", "Renewable"),
            "DSO",
            "ESB Networks",
            "MEC MW",
            "pdf",
        ),
        (
            "eirgrid_system_renewable_summary",
            first_matching_link(links_by_page.get("eirgrid_system_renewable", []), "System", "Renewable", "Summary"),
            "system",
            "EirGrid",
            "installed capacity / system summary",
            "xlsx",
        ),
    ]

    for source_key, url, network, publisher, basis, kind in jobs:
        if not url:
            audit["errors"].append({"source": source_key, "error": "No matching link discovered"})
            audit["sources"][source_key] = {"status": "missing_link"}
            continue

        try:
            if kind == "pdf":
                source_records, meta = parse_generator_pdf(
                    url,
                    source_key=source_key,
                    network=network,
                    publisher=publisher,
                    capacity_basis=basis,
                )
            elif kind == "xlsx":
                source_records, meta = scan_workbook_for_capacity(url, source_key)
            else:
                raise ValueError(f"Unsupported source kind: {kind}")

            records.extend(source_records)
            successful_source_keys.add(source_key)
            audit["sources"][source_key] = {
                "status": "loaded",
                **meta,
            }
        except Exception as exc:
            audit["errors"].append({"source": source_key, "url": url, "error": str(exc)})
            audit["sources"][source_key] = {
                "status": "failed",
                "url": url,
                "error": str(exc),
            }

    previous = load_previous()
    removals = removed_candidates(previous, records, successful_source_keys)
    records.extend(removals)

    deduped: dict[str, dict[str, Any]] = {}
    for rec in records:
        key = rec.get("id")
        if not key:
            continue
        if key not in deduped:
            deduped[key] = rec
        else:
            existing_sources = deduped[key].setdefault("sources", [])
            for src in rec.get("sources", []):
                if src not in existing_sources:
                    existing_sources.append(src)

    projects = sorted(
        deduped.values(),
        key=lambda r: (
            str(r.get("status", "")),
            str(r.get("technology", "")),
            str(r.get("name", "")).lower(),
        ),
    )

    evidence_sources = [
        {
            "key": "gov_ress",
            "publisher": "Department of Climate, Energy and the Environment",
            "status": audit["sources"].get("gov_ress", {}).get("status", "not_loaded"),
            "url": SOURCE_PAGES["gov_ress"],
            "evidence_role": "auction_support_pipeline",
        },
        {
            "key": "gov_oress",
            "publisher": "Department of Climate, Energy and the Environment",
            "status": audit["sources"].get("gov_oress", {}).get("status", "not_loaded"),
            "url": SOURCE_PAGES["gov_oress"],
            "evidence_role": "offshore_auction_pipeline",
        },
        {
            "key": "eirgrid_airaa",
            "publisher": "EirGrid and SONI",
            "status": audit["sources"].get("eirgrid_airaa", {}).get("status", "not_loaded"),
            "url": SOURCE_PAGES["eirgrid_airaa"],
            "evidence_role": "adequacy_and_retirement_assumptions",
        },
    ]

    model = {
        "meta": {
            "generated_at": now_iso(),
            "schema_version": "0.1.0",
            "builder": "ops/harvest_renewable_capacity.py",
            "mode": "official-register-harvest",
            "source_count": len(audit["sources"]),
            "successful_source_count": len([s for s in audit["sources"].values() if s.get("status") == "loaded"]),
            "error_count": len(audit["errors"]),
            "caveat": (
                "This dataset separates connected, contracted and support-awarded capacity. "
                "Removed projects are candidates only unless independently confirmed as closed, withdrawn or retired."
            ),
        },
        "summary": summarise(projects),
        "evidence_sources": evidence_sources,
        "projects": projects,
    }

    return model, audit


def main() -> int:
    model, audit = harvest()
    OUT.write_text(json.dumps(model, indent=2, ensure_ascii=False) + "\n")
    AUDIT_OUT.write_text(json.dumps(audit, indent=2, ensure_ascii=False) + "\n")
    print(f"Wrote {OUT.relative_to(ROOT)} with {len(model.get('projects', []))} records")
    print(f"Wrote {AUDIT_OUT.relative_to(ROOT)}")
    if audit.get("errors"):
        print(f"WARNING: {len(audit['errors'])} renewable-capacity source issue(s). See {AUDIT_OUT.relative_to(ROOT)}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
